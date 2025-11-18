import io
import json
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher

# ── Page meta / theme ────────────────────────────────────────────────
st.set_page_config(page_title="Masterfile Automation - Target", page_icon="🎯", layout="wide", initial_sidebar_state="collapsed")
st.markdown("""
<style>
:root{ --bg1:#f6f9fc; --bg2:#fff; --card:#fff; --card-border:#e8eef6; --ink:#0f172a; --muted:#64748b; --accent:#cc0000; }
.stApp{background:linear-gradient(180deg, var(--bg1) 0%, var(--bg2) 70%);}
.block-container{padding-top:.75rem; max_width: 1200px;}
.section{border:1px solid var(--card-border); background:var(--card); border-radius:16px; padding:18px 20px; box-shadow:0 6px 24px rgba(2,6,23,.05); margin-bottom:18px;}
.badge{display:inline-block; padding:4px 10px; border-radius:999px; font-size:.82rem; font-weight:600; margin-right:.25rem;}
.badge-info{background:#eef2ff;color:#1e40af;} .badge-ok{background:#ecfdf5;color:#065f46;} .badge-target{background:#fff0f0;color:#cc0000;}
div.stButton>button,.stDownloadButton>button{background:var(--accent)!important;color:#fff!important;border-radius:10px!important;border:0!important;font-weight:600!important;padding:0.5rem 1.5rem!important;transition:all .3s ease!important;}
div.stButton>button:hover,.stDownloadButton>button:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(204,0,0,0.3)!important;}
h1{color:var(--accent);}
</style>
""", unsafe_allow_html=True)

# ── Template constants ───────────────────────────────────────────────
MASTER_TEMPLATE_SHEET = "Bulk Product Data"
MASTER_DISPLAY_ROW    = 1
MASTER_SECONDARY_ROW  = 2
MASTER_DATA_START_ROW = 3

# ── XML namespaces ───────────────────────────────────────────────────
XL_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XL_NS_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", XL_NS_MAIN)
ET.register_namespace("r", XL_NS_REL)
ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")

# ── Helpers ─────────────────────────────────────────────────────────
_INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF]")
def sanitize_xml_text(s): return "" if s is None else _INVALID_XML_CHARS.sub("", str(s))
def norm(s: str) -> str:
    if s is None: return ""
    x = str(s).strip().lower()
    x = re.sub(r"\s*-\s*en\s*[-_ ]\s*us\s*$", "", x)
    x = x.replace("–","-").replace("—","-").replace("−","-")
    x = re.sub(r"[._/\\-]+", " ", x)
    x = re.sub(r"[^0-9a-z\s]+", " ", x)
    return re.sub(r"\s+", " ", x).strip()
def top_matches(query, candidates, k=3):
    q = norm(query)
    scored = [(SequenceMatcher(None, q, norm(c)).ratio(), c) for c in candidates]
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[:k]
def nonempty_rows(df: pd.DataFrame) -> int:
    if df.empty: return 0
    return df.replace("", pd.NA).dropna(how="all").shape[0]
def worksheet_used_cols(ws, header_rows=(1,), hard_cap=2048, empty_streak_stop=8):
    max_try = min(ws.max_column, hard_cap); last_nonempty=0; streak=0
    for c in range(1, max_try + 1):
        any_val = any((ws.cell(row=r, column=c).value not in (None, "")) for r in header_rows)
        if any_val: last_nonempty, streak = c, 0
        else:
            streak += 1
            if streak >= empty_streak_stop: break
    return max(last_nonempty, 1)
def _col_letter(n: int) -> str:
    s=""; 
    while n: n,r=divmod(n-1,26); s=chr(65+r)+s
    return s
def _col_number(letters: str) -> int:
    n=0
    for ch in letters:
        if not ch.isalpha(): break
        n = n*26 + (ord(ch.upper())-64)
    return n
def safe_filename(name: str, fallback: str = "final_masterfile"):
    if name is None: return fallback
    name = re.sub(r"[^A-Za-z0-9._ -]+", "", name.strip())
    return name or fallback

# ── Gender inference ─────────────────────────────────────────────────
_APOS = r"[’']"
_GENDER_W = re.compile(rf"\b(women(?:{_APOS}s)?|woman|female|lad(?:y|ies))\b", re.I)
_GENDER_M = re.compile(rf"\b(men(?:{_APOS}s)?|man|male|gent(?:lemen)?)\b", re.I)
_UNISEX   = re.compile(r"\b(unisex|all genders|everyone|for all|men\s*&\s*women|women\s*&\s*men)\b", re.I)
def _has_w(text: str) -> bool: return bool(_GENDER_W.search((text or "")))
def _has_m(text: str) -> bool: return bool(_GENDER_M.search((text or "")))
def _is_unisex(text: str) -> bool: return bool(_UNISEX.search((text or "").lower()))

# ── SEO field aliases (provided) ─────────────────────────────────────
SEO_ALIASES = {
    "Product Name": ["Product Name", "item_name", "Item Name", "Walmart Title - en-US", "Title"],
    "Description": ["Product Description", "Description", "long_description", "Walmart Description - en-US"],
    "bullet_point1": ["Bullet point 1","bullet_point1", "Bullet Feature 1", "bullet point 1", "bullet_point1 - en-US", "Key Features #1 - en-US"],
    "bullet_point2": ["Bullet point 2","bullet_point2", "Bullet Feature 2", "bullet point 2", "bullet_point2 - en-US", "Key Features #2 - en-US"],
    "bullet_point3": ["Bullet point 3","bullet_point3", "Bullet Feature 3", "bullet point 3", "bullet_point3 - en-US", "Key Features #3 - en-US"],
    "bullet_point4": ["Bullet point 4","bullet_point4", "Bullet Feature 4", "bullet point 4", "bullet_point4 - en-US", "Key Features #4 - en-US"],
    "bullet_point5": ["Bullet point 5","bullet_point5", "Bullet Feature 5", "bullet point 5", "bullet_point5 - en-US", "Key Features #5 - en-US"],
}
def select_seo_columns(df: pd.DataFrame) -> list[str]:
    header_lookup = {norm(c): c for c in df.columns}
    picks = []
    for _, aliases in SEO_ALIASES.items():
        for alias in aliases:
            key = norm(alias)
            if key in header_lookup:
                picks.append(header_lookup[key])
                break
    seen = set()
    picks = [c for c in picks if not (c in seen or seen.add(c))]
    if picks:
        return picks
    heur = [c for c in df.columns if any(k in norm(c) for k in ["title","product name","description","bullet","feature","name"])]
    return heur if heur else list(df.columns)
def _column_priority_score(col_name: str) -> int:
    n = norm(col_name)
    if "title" in n or "product name" in n: return 3
    if "bullet" in n or "feature" in n:     return 2
    if "description" in n:                   return 1
    return 0
def order_seo_columns(cols: list[str]) -> list[str]:
    return sorted(cols, key=_column_priority_score, reverse=True)

def infer_gender_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    for c in ordered_cols:
        t = str(row.get(c, ""))
        if _is_unisex(t):
            return "Gender Neutral"
    any_w = False; any_m = False
    for c in ordered_cols:
        t = str(row.get(c, ""))
        w = _has_w(t); m = _has_m(t)
        any_w = any_w or w
        any_m = any_m or m
        if w and not m:
            return "Women"
        if m and not w:
            return "Men"
    if any_w and any_m: return "Gender Neutral"
    if any_w: return "Women"
    if any_m: return "Men"
    return "Gender Neutral"

# ── Health & Beauty Subtype (≤3) ─────────────────────────────────────
_EXCLUDE_NON_POWDER = re.compile(r"\b(protein\s+bar|protein\s+cookie|protein\s+shake|ready[-\s]?to[-\s]?drink|rtd)\b", re.I)
_HB_SUBTYPE_PATTERNS = {
    "Collagen": [r"\bcollagen\b", r"\bcollagen\s+peptid(e|es)\b", r"\bhydrol(y|i)zed\s+collagen\b", r"\bmarine\s+collagen\b", r"\btype\s*(i|ii|iii)\b"],
    "Protein Powder": [r"\bprotein\s+powder\b", r"\bwhey\b", r"\bcasein\b", r"\bmicellar\s+casein\b", r"\b(isolate|concentrate)\b", r"\bpea\s+protein\b", r"\bsoy\s+protein\b", r"\brice\s+protein\b", r"\bprotein\s+blend\b"],
    "Multivitamins": [r"\bmult(i|i-)?vitamin(s)?\b", r"\bdaily\s+multivitamin(s)?\b", r"\bmulti[-\s]?vit\b"],
    "Vitamin A": [r"\bvit(amin)?\s*a\b", r"\bretinol\b", r"\bretinyl\b"],
    "Vitamin B": [r"\bvit(amin)?\s*b(\d{1,2})?\b", r"\bb[-\s]?complex\b", r"\bthiamin(e)?\b", r"\briboflavin\b", r"\bniacin(amide)?\b", r"\bpantothenic\b", r"\bpyridoxin(e)?\b", r"\bbiotin\b", r"\bfolate\b", r"\bfolic\s+acid\b", r"\bcobalamin\b", r"\bB-?12\b", r"\bB-?6\b", r"\bB-?3\b"],
    "Vitamin C": [r"\bvit(amin)?\s*c\b", r"\bascorb(ic|ate)\b", r"\bester[-\s]?c\b"],
    "Vitamin D": [r"\bvit(amin)?\s*d\b", r"\bd-?3\b", r"\bd-?2\b", r"\bcholecalciferol\b", r"\bergocalciferol\b"],
    "Vitamin E": [r"\bvit(amin)?\s*e\b", r"\btocopherol\b", r"\btocotrienol\b"],
    "Vitamin K": [r"\bvit(amin)?\s*k\b", r"\bk-?2\b", r"\bmk-?\s?7\b", r"\bmenaquinone\b", r"\bphylloquinone\b"],
}
def _hb_score_patterns(text: str, patterns: list[str]) -> int:
    if not text: return 0
    return sum(1 for p in patterns if re.search(p, text, re.I))
def infer_hb_subtype_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    scores = {k: 0 for k in _HB_SUBTYPE_PATTERNS.keys()}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt:
            continue
        weight = _column_priority_score(c)
        protein_powder_ok = True
        if _EXCLUDE_NON_POWDER.search(txt) and not re.search(r"\bprotein\s+powder\b", txt, re.I):
            protein_powder_ok = False
        for label, pats in _HB_SUBTYPE_PATTERNS.items():
            if label == "Protein Powder" and not protein_powder_ok:
                continue
            hits = _hb_score_patterns(txt, pats)
            if hits:
                scores[label] += weight * hits
    picks = [(k, v) for k, v in scores.items() if v > 0]
    if not picks:
        return ""
    picks.sort(key=lambda kv: (-kv[1], kv[0]))
    return "|".join([k for k, _ in picks[:3]])

# ── Health Application* (≤5) ─────────────────────────────────────────
_HEALTH_APP_LABELS = [
    "Adrenal Health","Aging","Allergies","Anxiety","Bladder Infection","Bladder Support","Bloating","Blood Clots",
    "Blood Sugar Imbalance","Bone Health","Children's Health","Cholesterol Level Maintenance","Circulatory System Health",
    "Constipation","Dental Health","Diabetes","Diarrhea","Digestive Health","Endurance","Energy","eye health","Fertility",
    "Fever","Gout","Hair, Skin and Nail Health","Heart Health","High Cholesterol","Homocysteine Levels","Hydration",
    "Immune System Health","Infection","Inflammation","Insomnia","Intestinal Health","Iron Deficiency",
    "Irritable Bowel Syndrome (IBS)","Joint Health","Joint Pain","Joint Support","Kidney Health","Lactation",
    "Liver Health","Lymphatic Health","Memory and Brain Health","Men's Health","Menopause","Metabolism","Mood",
    "Morning Sickness","Muscle Growth","Muscle Pain","Muscle Tension","Nail Health","Nausea","Nerve Pain",
    "Nervous System Health","overall health","Pain Relief","PMS","Postnatal Health","Postpartum Care","Pregnancy",
    "Premenstrual Breast Discomfort","Prenatal Health","Pressure Ulcers","Prostate Health","Respiratory Health",
    "Seasonal Allergies","Sexual Health","Sinusitis","Skin Health","Sleep Disturbance","Sleep Support",
    "Sports Performance","Strength","Stress","Testosterone Level","Thyroid Health","Tinnitus","Uric Acid Levels",
    "Urinary Health","Urinary Tract Infection","Vaginal Health","Vertigo","Water Retention","Weight Loss",
    "Weight Management","Women's Health","Yeast Infection"
]
def _make_base_pat(label: str) -> str:
    s = label.lower()
    s = s.replace("children's", r"children(?:'s)?").replace("men's", r"men(?:'s)?").replace("women's", r"women(?:'s)?")
    s = re.sub(r"\s*\(.*?\)\s*", "", s)
    tokens = re.split(r"[^a-z0-9]+", s.strip())
    tokens = [re.escape(t) for t in tokens if t]
    if not tokens:
        return r"$^"
    return r"\b" + r"\W+".join(tokens) + r"\b"
_HEALTH_APP_SYNONYMS = {
    "Digestive Health": [r"\bdigesti(ve|on)\b", r"\bgut\s+health\b"],
    "Immune System Health": [r"\bimmune(\s+system)?\s+health\b", r"\bimmune\s+support\b", r"\bimmunity\s+support\b", r"\bboost\s+immun"],
    "Joint Health": [r"\bjoint\s+health\b", r"\bhealthy\s+joints\b"],
    "Joint Support": [r"\bjoint\s+support\b"],
    "Joint Pain": [r"\bjoint\s+pain\b", r"\barthriti[cs]\b", r"\barthritic\s+pain\b"],
    "Sleep Support": [r"\bsleep\s+support\b", r"\bbetter\s+sleep\b", r"\bpromotes\s+sleep\b", r"\bsleep\s+quality\b"],
    "Sleep Disturbance": [r"\bsleep\s+disturbance\b", r"\btrouble\s+sleep(ing)?\b"],
    "Insomnia": [r"\binsomnia\b"],
    "Energy": [r"\benergy\b", r"\benergiz", r"\bpre[-\s]?workout\b"],
    "Endurance": [r"\bendurance\b", r"\bstamina\b"],
    "Stress": [r"\bstress\b", r"\bstress\s+relief\b", r"\bstress\s+support\b"],
    "Anxiety": [r"\banxiety\b"],
    "Weight Loss": [r"\bweight\s+loss\b", r"\bfat\s+loss\b", r"\bslim(ming)?\b"],
    "Weight Management": [r"\bweight\s+management\b", r"\bmanage\s+weight\b", r"\bmaintain\s+weight\b"],
    "Bone Health": [r"\bbone\s+health\b", r"\bbone\s+density\b", r"\bosteoporosis\b"],
    "Heart Health": [r"\bheart\s+health\b", r"\bcardio(vascular)?\b"],
    "Skin Health": [r"\bskin\s+health\b", r"\bhealthy\s+skin\b"],
    "Hair, Skin and Nail Health": [r"\bhair[, ]+\s*skin\s*(and|&)?\s*nail", r"\bhair\s*skin\s*and\s*nails\b"],
    "Memory and Brain Health": [r"\bmemory\b", r"\bbrain\s+health\b", r"\bcognit(ive|ion)\b", r"\bfocus\b", r"\bconcentration\b", r"\bnootropic\b"],
    "Hydration": [r"\bhydrat(e|ion)\b", r"\belectrolyte(s)?\b"],
    "Cholesterol Level Maintenance": [r"\bcholesterol\b", r"\bmaintain\s+cholesterol\b", r"\bhealthy\s+cholesterol\b"],
    "High Cholesterol": [r"\bhigh\s+cholesterol\b"],
    "Blood Sugar Imbalance": [r"\bblood\s+sugar\b", r"\bglucose\b", r"\bglycem(i|y)c\b"],
    "Children's Health": [r"\b(children|kids|child)\b.*\bhealth\b", r"\bfor\s+(kids|children)\b"],
    "Men's Health": [r"\b(men|male)\b.*\bhealth\b", r"\bfor\s+men\b"],
    "Women's Health": [r"\b(women|female)\b.*\bhealth\b", r"\bfor\s+women\b"],
    "Irritable Bowel Syndrome (IBS)": [r"\birritable\s+bowel\s+syndrome\b", r"\bIBS\b"],
    "eye health": [r"\beye\s+health\b", r"\bvision\b", r"\bocular\b"],
    "Nervous System Health": [r"\bnervous\s+system\b", r"\bneurolog(y|ical)\b"],
    "Respiratory Health": [r"\brespiratory\b", r"\blung\b", r"\bbreath(ing)?\b"],
    "Sports Performance": [r"\bsports?\s+performance\b", r"\bathletic\b", r"\bperformance\b"],
    "Strength": [r"\bstrength\b", r"\bstronger\b"],
    "Pain Relief": [r"\bpain\s+relief\b", r"\banalgesic\b"],
    "Mood": [r"\bmood\b"],
    "Metabolism": [r"\bmetaboli[sc]m\b"],
}
_HEALTH_APP_REGEX = {}
for label in _HEALTH_APP_LABELS:
    pats = [ _make_base_pat(label) ]
    pats.extend(_HEALTH_APP_SYNONYMS.get(label, []))
    _HEALTH_APP_REGEX[label] = [re.compile(p, re.I) for p in pats]
def _health_score(text: str, comp_list) -> int:
    if not text: return 0
    return sum(1 for rx in comp_list if rx.search(text))
def infer_health_app_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    scores = {label:0 for label in _HEALTH_APP_REGEX.keys()}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt: 
            continue
        weight = _column_priority_score(c)
        for label, comp_list in _HEALTH_APP_REGEX.items():
            hits = _health_score(txt, comp_list)
            if hits:
                scores[label] += weight * hits
    picks = [(k, v) for k, v in scores.items() if v > 0]
    if not picks:
        return ""
    picks.sort(key=lambda kv: (-kv[1], kv[0]))
    return "|".join([k for k, _ in picks[:5]])

# ── Targeted Audience* (single; default Adult) ───────────────────────
_AUD_PAT = {
    "Infant": [r"\bbaby|babies|infant|newborn\b", r"\b0\s*[-–]?\s*12\s*(m|mos|months)\b"],
    "Kids":   [r"\bkid(s)?\b", r"\bchild(ren)?\b", r"\btoddler(s)?\b", r"\bpre[-\s]?school\b"],
    "Teen":   [r"\bteen(s|age|ager|agers)?\b", r"\byouth\b"],
    "Adult":  [r"\badult(s)?\b"]
}
_AGE_YEARS_RX = re.compile(r"\b(\d{1,2})\s*(?:\+|plus)?\s*(?:y(?:rs?)?|years?)\b", re.I)
_AGE_RANGE_YEARS_RX = re.compile(r"\b(\d{1,2})\s*[-–]\s*(\d{1,2})\s*(?:y(?:rs?)?|years?)\b", re.I)
_AGE_MONTHS_RX = re.compile(r"\b(\d{1,2})\s*(?:m|mos|months?)\b", re.I)
def _aud_bump(scores: dict, label: str, w: int = 1):
    scores[label] = scores.get(label, 0) + w
def _age_to_bucket(years: int | None = None, months: int | None = None) -> str | None:
    if months is not None:
        if months <= 24: return "Infant"
        elif months <= 36: return "Kids"
        else: return "Kids"
    if years is not None:
        if years <= 2: return "Infant"
        if 3 <= years <= 12: return "Kids"
        if 13 <= years <= 17: return "Teen"
        if years >= 18: return "Adult"
    return None
def infer_targeted_audience(row: pd.Series, ordered_cols: list[str], gender_val: str = "") -> str:
    scores = {"Adult":0, "Infant":0, "Kids":0, "Teen":0}
    if str(gender_val).strip() in ("Men","Women"):
        _aud_bump(scores, "Adult", 2)
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt: 
            continue
        w = _column_priority_score(c)
        for label, pats in _AUD_PAT.items():
            for p in pats:
                if re.search(p, txt, re.I):
                    _aud_bump(scores, label, w)
        for m in _AGE_YEARS_RX.finditer(txt):
            y = int(m.group(1))
            bucket = _age_to_bucket(years=y)
            if bucket: _aud_bump(scores, bucket, w+1)
        for m in _AGE_RANGE_YEARS_RX.finditer(txt):
            y1, y2 = int(m.group(1)), int(m.group(2))
            for y in (y1, y2):
                bucket = _age_to_bucket(years=y)
                if bucket: _aud_bump(scores, bucket, w+1)
        for m in _AGE_MONTHS_RX.finditer(txt):
            mo = int(m.group(1))
            bucket = _age_to_bucket(months=mo)
            if bucket: _aud_bump(scores, bucket, w+1)
    order = ["Infant","Kids","Teen","Adult"]
    best = max(order, key=lambda lab: (scores.get(lab,0), -order.index(lab)))
    if scores.get(best,0) == 0:
        return "Adult"
    return best

# ── Product Form* (single; 'Multiple Forms' if >1 strong) ────────────
_PRODUCT_FORM_PATTERNS = {
    "Bar":               [r"\b(protein|nutrition)\s+bar\b", r"\bbar\b"],
    "Caplet":            [r"\bcaplet(s)?\b"],
    "Capsule":           [r"\bcapsule(s)?\b", r"\bcaps?\b", r"\bveg(?:etable)?\s*caps?(?:ule)?s?\b"],
    "Chewable Tablet":   [r"\bchewable\s+tablet(s)?\b", r"\bODT\b", r"\borally\s+disintegrating\s+tablet(s)?\b", r"\bfast[-\s]?dissolv(e|ing)\s+tablet(s)?\b"],
    "Chewable":          [r"\bchewable\b"],
    "Cream":             [r"\bcream(s)?\b"],
    "Dissolving Strip":  [r"\bdissolving\s+strip(s)?\b", r"\boral\s+strip(s)?\b", r"\bmouth\s+strip(s)?\b"],
    "Dissolving Tablet": [r"\bdissolving\s+tablet(s)?\b", r"\beffervescent\s+tablet(s)?\b", r"\bsublingual\s+tablet(s)?\b"],
    "Gel":               [r"\btopical\s+gel\b", r"\bgel\b"],
    "Gelcap":            [r"\bgel[-\s]?cap(s)?\b", r"\bgelcap(s)?\b"],
    "Gum":               [r"\bgum\b"],
    "Gummy":             [r"\bgummy\b", r"\bgummies\b"],
    "Liquid":            [r"\bliquid\b", r"\bsyrup\b", r"\bdrops?\b", r"\belixir\b", r"\btincture\b", r"\bsuspension\b", r"\bsolution\b"],
    "Lollipop":          [r"\blollipop(s)?\b"],
    "Lozenge":           [r"\blozenge(s)?\b", r"\bpastille(s)?\b", r"\btroche(s)?\b", r"\bthroat\s+drops?\b", r"\bcough\s+drops?\b"],
    "Patch":             [r"\bpatch(es)?\b", r"\btransdermal\b"],
    "Powder":            [r"\bpowder(ed)?\b", r"\bdrink\s+mix\b"],
    "Softgel":           [r"\bsoft[-\s]?gel(s)?\b", r"\bsoftgel(s)?\b"],
    "Tablet":            [r"\btablet(s)?\b", r"\btabs?\b"],
    "Tea":               [r"\btea\b(?!\s*tree)"],
    "Wafer":             [r"\bwafer(s)?\b"],
}
_PF_EXCLUDE = {
    "Gel": [r"\bsoft[-\s]?gel(s)?\b", r"\bgel[-\s]?cap(s)?\b", r"\bgelcap(s)?\b"],
    "Capsule": [r"\bsoft[-\s]?gel(s)?\b", r"\bgelcap(s)?\b", r"\bgel[-\s]?cap(s)?\b"],
    "Chewable": [r"\bchewable\s+tablet(s)?\b", r"\bgummies?\b"],
    "Liquid": [r"\bliquid\s+softgel(s)?\b"],
    "Tablet": [r"\bchewable\s+tablet(s)?\b", r"\bdissolving\s+tablet(s)?\b", r"\beffervescent\s+tablet(s)?\b", r"\bsublingual\s+tablet(s)?\b"],
    "Gum": [r"\bgummies?\b"],
    "Tea": [r"\btea\s*tree\b"],
}
def _match_any(rx_list, text) -> int:
    if not text: return 0
    return sum(1 for p in rx_list if re.search(p, text, re.I))
def _excluded(label: str, text: str) -> bool:
    for p in _PF_EXCLUDE.get(label, []):
        if re.search(p, text, re.I):
            return True
    return False
def infer_product_form_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    scores = {k:0 for k in _PRODUCT_FORM_PATTERNS.keys()}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt:
            continue
        w = _column_priority_score(c)
        for label, pats in _PRODUCT_FORM_PATTERNS.items():
            if _excluded(label, txt):
                continue
            hits = _match_any(pats, txt)
            if hits:
                scores[label] += w * hits
    if scores["Chewable Tablet"] > 0:
        scores["Chewable"] = 0
        scores["Tablet"] = max(0, scores["Tablet"] - 1)
    if scores["Dissolving Tablet"] > 0:
        scores["Tablet"] = 0
    if scores["Softgel"] > 0 or scores["Gelcap"] > 0:
        scores["Capsule"] = 0
        scores["Gel"] = max(0, scores["Gel"] - 1)
    if scores["Gummy"] > 0:
        scores["Gum"] = 0
    picks = [(k, v) for k, v in scores.items() if v > 0]
    if not picks:
        return ""
    picks.sort(key=lambda kv: (-kv[1], kv[0]))
    top_label, top_score = picks[0]
    other_score = sum(v for _, v in picks[1:])
    if len(picks) >= 2 and top_score < other_score * 1.5:
        return "Multiple Forms"
    return top_label

# ── primary flavors (≤3; pipe-delimited) ────────────────────────────
_FLAVOR_LABELS = [
    "Acai","Almond","Apple","Banana","Berry","Black Cherry","Black Currant","Black Tea","Blackberry","Blue Raspberry",
    "Blueberry","Brown Sugar","Brownie","Bubble Gum","Butter","Cake","Caramel","Celery","Chai","Chai Latte","Chamomile",
    "Cherry","Chocolate","Chocolate Chip","Cinnamon","Citrus","Cocoa","Coconut","Coffee","Cookie","Cookies and Cream",
    "Cranberry","Dark Chocolate","Donut","Dragon Fruit","Elderberry","Flavored","Flaxseed","French Vanilla","Fresh",
    "Fruit","Fruit Punch","Garlic","Goji Berry","Grape","Grapefruit","Green Apple","Green Tea","Guava","Hibiscus","Honey",
    "Kiwi","Lemon","Lemonade","Lime","Macadamia Nut","Mango","Maple","Marshmallow","Matcha","Melon","Milk","Milk Chocolate",
    "Mint","Mixed Berry","Mixed Fruit","Mocha","Mushroom","Natural","No Flavor","Nut","Oatmeal","Oats","Orange","Passion Fruit",
    "Peach","Peanut Butter","Pear","Peppermint","Pineapple","Pistachio","Pomegranate","Raspberry","Salted Caramel","Seaberry",
    "Sour","Spicy","Spirulina","Strawberry","Sugar","Tangerine","Tea","Toffee","Tropical Fruit","Turmeric","Ube","Unflavored",
    "Vanilla","Vegetable Blend","Watermelon","White Chocolate","Whole Wheat","Wild Berry","Yuzu"
]
_FLAVOR_PAT = {
    "Blue Raspberry": [r"\bblue\s+rasp(berry)?\b", r"\bblue\s*razz\b", r"\bbluerazz\b"],
    "Cookies and Cream": [r"cookies?\s*(and|&|n)\s*cream", r"cookie\s*n\s*cream"],
    "French Vanilla": [r"\bfrench\s+vanilla\b"],
    "Milk Chocolate": [r"\bmilk\s+choc(olate)?\b"],
    "Dark Chocolate": [r"\bdark\s+choc(olate)?\b"],
    "White Chocolate": [r"\bwhite\s+choc(olate)?\b"],
    "Salted Caramel": [r"\bsalted\s+caramel\b"],
    "Fruit Punch": [r"\bfruit\s+punch\b"],
    "Mixed Berry": [r"\bmixed\s+berr(y|ies)\b", r"\bberry\s+blend\b"],
    "Wild Berry": [r"\bwild\s+berry\b", r"\bwildberry\b"],
    "Green Tea": [r"\bgreen\s+tea\b"],
    "Black Tea": [r"\bblack\s+tea\b"],
    "Chai Latte": [r"\bchai\s+latte\b"],
    "Matcha": [r"\bmatcha\b"],
    "Unflavored": [r"\bunflavor(ed)?\b", r"\bno\s*flavor\b", r"\bplain\b", r"\boriginal\b"],
}
for lab in _FLAVOR_LABELS:
    if lab not in _FLAVOR_PAT:
        _FLAVOR_PAT[lab] = [rf"\b{re.escape(lab.lower())}\b"]
_FLAVOR_DEMOTE_IF_CHILD = {
    "Chocolate": ["Milk Chocolate","Dark Chocolate","White Chocolate","Chocolate Chip"],
    "Vanilla": ["French Vanilla"],
    "Tea": ["Green Tea","Black Tea","Chai","Chai Latte","Chamomile","Hibiscus","Matcha"],
    "Berry": ["Blueberry","Raspberry","Strawberry","Blackberry","Elderberry","Goji Berry","Seaberry","Mixed Berry","Wild Berry"],
    "Fruit": ["Mixed Fruit","Tropical Fruit","Fruit Punch","Orange","Apple","Mango","Peach","Grape","Guava","Pineapple","Watermelon","Pomegranate","Dragon Fruit","Passion Fruit","Tangerine","Grapefruit","Kiwi","Pear","Lemon","Lime","Green Apple"],
    "Caramel": ["Salted Caramel"],
    "Natural": ["Unflavored"],
    "No Flavor": ["Unflavored"]
}
_FLAVOR_LOW_PRIORITY = {"Flavored","Fresh","Natural","Fruit","Tea","Berry","Milk","Nut","Sugar"}
def _flavor_hits(label: str, text: str) -> int:
    pats = _FLAVOR_PAT.get(label, [])
    return sum(1 for p in pats if re.search(p, text or "", re.I))
def infer_primary_flavors_from_columns(row: pd.Series, ordered_cols: list[str], max_picks: int = 3) -> str:
    scores = {lab: 0.0 for lab in _FLAVOR_LABELS}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt:
            continue
        w = float(_column_priority_score(c) or 1)
        for lab in _FLAVOR_LABELS:
            hits = _flavor_hits(lab, txt)
            if hits:
                base = hits
                if " " in lab:
                    base += 0.5
                if lab in _FLAVOR_LOW_PRIORITY:
                    base -= 0.25
                scores[lab] += w * base
    present = {lab for lab, sc in scores.items() if sc > 0}
    for parent, children in _FLAVOR_DEMOTE_IF_CHILD.items():
        if parent in present and any(ch in present for ch in children):
            scores[parent] *= 0.25
    picks = [(k, v) for k, v in scores.items() if v >= 1.5]
    if not picks:
        if scores.get("Unflavored", 0) > 0:
            return "Unflavored"
        return ""
    picks.sort(key=lambda kv: (-kv[1], kv[0]))
    selected = [k for k, _ in picks[:max_picks]]
    if "Unflavored" in selected:
        selected = [s for s in selected if s not in ("No Flavor","Natural")]
    return "|".join(selected)

# ── Food & Drink Form 1 (single) ─────────────────────────────────────
_FD_LIQUID_EXCLUDE = re.compile(
    r"\b(soft[-\s]?gel|gelcap|capsule|tablet|pill|shampoo|conditioner|soap|detergent|cleaner|sanitizer|serum|lotion|toner)\b",
    re.I
)
_FD_PATTERNS = {
    "Granulated": [r"\bgranulated\b", r"\bgranular\b", r"\bgranule(s)?\b"],
    "Powdered": [r"\bpowder(ed)?\b", r"\bpowder\s+mix\b", r"\bdrink\s+mix\b", r"\binstant\s+powder\b"],
    "Liquid Concentrate": [
        r"\bliquid\s+concentrate\b",
        r"\bconcentrated\s+(?:drink|beverage|syrup|juice)\b",
        r"\bwater\s+enhancer\b",
        r"\b(drink|flavor|beverage)\s+concentrate\b",
        r"\b(squash|cordial)\b"
    ],
    "Liquid": [r"\bliquid\b", r"\bsyrup\b", r"\bready[-\s]?to[-\s]?drink\b", r"\brtd\b"],
}
def _fd_hits(patterns: list[str], text: str) -> int:
    if not text: return 0
    return sum(1 for p in patterns if re.search(p, text, re.I))
def infer_food_and_drink_form1_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    scores = {k: 0.0 for k in _FD_PATTERNS.keys()}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt:
            continue
        w = float(_column_priority_score(c) or 1)
        is_non_food_liquid = bool(_FD_LIQUID_EXCLUDE.search(txt))
        for label, pats in _FD_PATTERNS.items():
            hits = _fd_hits(pats, txt)
            if not hits:
                continue
            score = w * hits
            if label == "Liquid" and is_non_food_liquid:
                score *= 0.1
            scores[label] += score
    if scores["Liquid Concentrate"] > 0:
        scores["Liquid"] *= 0.2
    best_label, best_score = max(scores.items(), key=lambda kv: kv[1])
    return best_label if best_score > 0 else ""

# ── Tax* inference (single best match) ───────────────────────────────
_TAX_LABELS = [
    "Baby Monitors w/screen size >=4\" and < 15\"",
    "E-Bikes-Trikes_<1-HP_(750W)_Has-Pedals",
    "Paint 2 PK - (> 8 Ounces < 1 Gallon)",
    "Portable Gas Cans <5 Gallons",
    # — SNIP — keep your full list here —
    "Window and Door Weatherization","Work Gloves","Yarn, Elastic, Thread, Buttons"
]
_TAX_EXCLUDE = {"General"}
_STOP_TOKENS = {"and","or","with","without","for","the","a","an","of","to","in","on","by","pk","pack","pcs","pc",
                "oz","ounce","ounces","lb","lbs","g","gram","grams","ml","l","liter","liters","size","screen",
                "single","dual","us","less","more","than","equal","inch","inches","gal","gallon"}
def _normalize_token(tok: str) -> str:
    t = tok.lower()
    if t.endswith("ies"): t = t[:-3] + "y"
    elif t.endswith("es"): t = t[:-2]
    elif t.endswith("s") and len(t) > 3: t = t[:-1]
    repl = {"children":"child","babies":"baby","women":"woman","men":"man","webcam":"web camera"}
    return repl.get(t, t)
def _tokens(s: str) -> set[str]:
    raw = re.findall(r"[a-z0-9]+", (s or "").lower())
    toks = set()
    for t in raw:
        n = _normalize_token(t)
        if n and n not in _STOP_TOKENS:
            toks.add(n)
    return toks
_TAX_LABEL_TOKENS = {lab: _tokens(lab) for lab in _TAX_LABELS}
_TAX_SYNONYMS = {
    "Deodorant": [r"\bdeodorant(s)?\b"],
    "Antiperspirant": [r"\banti[-\s]?perspirant(s)?\b"],
    "Sunscreen": [r"\bsun\s*screen\b", r"\bspf\b"],
    "Baby Wipes": [r"\b(baby|infant)\s+wipes?\b"],
    "Diaper Bags": [r"\bdiaper\s+bag(s)?\b"],
    "Strollers": [r"\bstroller(s)?\b", r"\btravel\s+system(s)?\b"],
    "Earbuds, Headphones and Web Cameras": [r"\bearbud(s)?\b", r"\bheadphone(s)?\b", r"\bheadset(s)?\b", r"\bweb\s*cam(s)?\b"],
    "LED-Light Bulb": [r"\bled\b.*\blight\b|\bled\b.*\bbulb\b"],
    "CFL Light Bulb 2 PK": [r"\bcfl\b.*\b(light|bulb)\b"],
    "Incandescent-Light Bulb": [r"\bincandescent\b.*\b(light|bulb)\b"],
    "Soft Drinks": [r"\bsoft\s+drink(s)?\b", r"\bsoda(s)?\b", r"\bpop\b"],
    "Coffee Drinks  - With Milk": [r"\bcoffee\b.*\b(milk|latte|cappuccino)\b"],
    "Coffee Drinks - Unsweetened": [r"\bblack\s+coffee\b"],
    "Candy": [r"\bcandy\b|chocolate|toffee|taffy|caramel"],
    "Gum": [r"\bgum\b|chewing\s+gum"],
    "Water Bottles": [r"\bwater\s+bottle(s)?\b"],
    "Backpacks & Book Bags": [r"\bback\s*pack(s)?\b|\bbook\s*bag(s)?\b"],
    "Batteries": [r"\b(aa|aaa|c|d|9v)\b.*\bbatter(y|ies)\b|\bbatter(y|ies)\b"],
}
def _syn_hits(label: str, text: str) -> int:
    pats = _TAX_SYNONYMS.get(label, [])
    return sum(1 for p in pats if re.search(p, text or "", re.I))
def infer_tax_from_columns(row: pd.Series, ordered_cols: list[str]) -> str:
    scores = {lab: 0.0 for lab in _TAX_LABELS}
    for c in ordered_cols:
        txt = str(row.get(c, "")) or ""
        if not txt:
            continue
        w = float(_column_priority_score(c) or 1)
        text_tokens = _tokens(txt)
        text_norm = " " + norm(txt) + " "
        for lab, lab_tokens in _TAX_LABEL_TOKENS.items():
            if lab in _TAX_EXCLUDE:
                continue
            inter = lab_tokens.intersection(text_tokens)
            token_score = len(inter)
            phrase_boost = 0.0
            if token_score >= 2 and all(t in text_norm for t in lab_tokens):
                phrase_boost = 1.5
            syn_boost = 0.75 * _syn_hits(lab, txt)
            scores[lab] += w * (token_score + phrase_boost + syn_boost)
    best_label, best_score = "", 0.0
    for lab, sc in scores.items():
        if sc > best_score:
            best_label, best_score = lab, sc
    if best_score <= 2.0:
        return ""
    return best_label

# ── ZIP / XML helpers ────────────────────────────────────────────────
def _find_sheet_part_path(z: zipfile.ZipFile, sheet_name: str) -> str:
    wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid=None
    for sh in wb_xml.find(f"{{{XL_NS_MAIN}}}sheets"):
        if sh.attrib.get("name")==sheet_name:
            rid=sh.attrib.get(f"{{{XL_NS_REL}}}id"); break
    if not rid: raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    target=None
    for rel in rels_xml:
        if rel.attrib.get("Id")==rid:
            target=rel.attrib.get("Target"); break
    if not target: raise ValueError(f"Relationship for sheet '{sheet_name}' not found.")
    target = target.replace("\\","/")
    if target.startswith("../"): target=target[3:]
    if not target.startswith("xl/"): target="xl/"+target
    return target

def _get_table_paths_for_sheet(z: zipfile.ZipFile, sheet_path: str) -> list:
    rels_path = sheet_path.replace("worksheets/","worksheets/_rels/").replace(".xml",".xml.rels")
    if rels_path not in z.namelist(): return []
    root = ET.fromstring(z.read(rels_path)); out=[]
    for rel in root:
        if rel.attrib.get("Type","").endswith("/table"):
            target = rel.attrib.get("Target","").replace("\\","/")
            if target.startswith("../"): target=target[3:]
            if not target.startswith("xl/"): target="xl/"+target
            out.append(target)
    return out

def _read_table_cols_count(table_xml_bytes: bytes) -> int:
    try:
        root = ET.fromstring(table_xml_bytes)
        tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
        if tcols is None: return 0
        cnt_attr = tcols.attrib.get("count")
        cnt = int(cnt_attr) if cnt_attr else 0
        child_count = sum(1 for _ in tcols)
        return max(cnt, child_count)
    except Exception:
        return 0

def _union_dimension(orig_dim_ref: str, used_cols: int, last_row: int) -> str:
    try:
        _, right = orig_dim_ref.split(":", 1)
        m = re.match(r"([A-Z]+)(\d+)", right)
        if m:
            orig_last_col=_col_number(m.group(1)); orig_last_row=int(m.group(2))
        else:
            orig_last_col,orig_last_row=used_cols,last_row
    except Exception:
        orig_last_col,orig_last_row=used_cols,last_row
    u_last_col=max(orig_last_col,used_cols); u_last_row=max(orig_last_row,last_row)
    return f"A1:{_col_letter(u_last_col)}{u_last_row}"

def _ensure_ws_x14ac(root):
    root.set("{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable","x14ac")

def _intersects_range(a1: str, r1: int, r2: int) -> bool:
    m = re.match(r"^[A-Z]+(\d+):[A-Z]+(\d+)$", a1 or "", re.I)
    if not m: return False
    lo=int(m.group(1)); hi=int(m.group(2))
    if lo>hi: lo,hi=hi,lo
    return not (hi<r1 or lo>r2)

# ── Writer (inlineStr only) ──────────────────────────────────────────
def _patch_sheet_xml(sheet_xml_bytes: bytes, header_row: int, start_row: int, used_cols_final: int, block_2d: list) -> bytes:
    root = ET.fromstring(sheet_xml_bytes)
    _ensure_ws_x14ac(root)
    sheetData = root.find(f"{{{XL_NS_MAIN}}}sheetData")
    if sheetData is None:
        sheetData = ET.SubElement(root, f"{{{XL_NS_MAIN}}}sheetData")
    for row in list(sheetData):
        try: r = int(row.attrib.get("r") or "0")
        except Exception: r = 0
        if r >= start_row:
            sheetData.remove(row)
    mergeCells = root.find(f"{{{XL_NS_MAIN}}}mergeCells")
    if mergeCells is not None:
        for mc in list(mergeCells):
            if _intersects_range(mc.attrib.get("ref",""), start_row, 1048576):
                mergeCells.remove(mc)
        if len(list(mergeCells)) == 0:
            root.remove(mergeCells)
    row_span = f"1:{used_cols_final}" if used_cols_final > 0 else "1:1"
    n_rows = len(block_2d)
    for i in range(n_rows):
        r = start_row + i
        src_row = block_2d[i]
        row_el = ET.Element(f"{{{XL_NS_MAIN}}}row", r=str(r))
        row_el.set("spans", row_span)
        row_el.set("{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent", "0.25")
        for j in range(used_cols_final):
            val = src_row[j] if j < len(src_row) else ""
            txt = sanitize_xml_text(val).strip() if val else ""
            if not txt:
                continue
            col = _col_letter(j + 1)
            c = ET.Element(f"{{{XL_NS_MAIN}}}c", r=f"{col}{r}", t="inlineStr")
            is_el = ET.SubElement(c, f"{{{XL_NS_MAIN}}}is")
            t_el = ET.SubElement(is_el, f"{{{XL_NS_MAIN}}}t")
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            t_el.text = txt
            row_el.append(c)
        sheetData.append(row_el)
    dim = root.find(f"{{{XL_NS_MAIN}}}dimension")
    if dim is None:
        dim = ET.SubElement(root, f"{{{XL_NS_MAIN}}}dimension", ref="A1:A1")
    last_row = max(header_row, start_row + max(0, n_rows) - 1)
    dim.set("ref", _union_dimension(dim.attrib.get("ref", "A1:A1"), used_cols_final, last_row))
    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter")
    if af is not None:
        af.set("ref", f"A{header_row}:{_col_letter(used_cols_final)}{last_row}")
    sheetPr = root.find(f"{{{XL_NS_MAIN}}}sheetPr")
    if sheetPr is not None and sheetPr.attrib.get("filterMode"):
        sheetPr.attrib.pop("filterMode", None)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _patch_table_xml(table_xml_bytes: bytes, header_row: int, last_row: int, last_col_n: int) -> bytes:
    root = ET.fromstring(table_xml_bytes)
    new_ref = f"A{header_row}:{_col_letter(last_col_n)}{last_row}"
    root.set("ref", new_ref)
    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter") or ET.SubElement(root, f"{{{XL_NS_MAIN}}}autoFilter")
    af.set("ref", new_ref)
    tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
    if tcols is not None:
        tcols.set("count", str(sum(1 for _ in tcols)))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _strip_calcchain_override(ct_bytes: bytes) -> bytes:
    try:
        ns="http://schemas.openxmlformats.org/package/2006/content-types"
        root=ET.fromstring(ct_bytes); ET.register_namespace("", ns)
        for el in list(root):
            if el.tag==f"{{{ns}}}Override" and el.attrib.get("PartName","").lower()=="/xl/calcchain.xml":
                root.remove(el)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)
    except Exception:
        return ct_bytes

def fast_patch_template(master_bytes: bytes, sheet_name: str, header_row: int, start_row: int, used_cols: int, block_2d: list) -> bytes:
    zin = zipfile.ZipFile(io.BytesIO(master_bytes), "r")
    sheet_path = _find_sheet_part_path(zin, sheet_name)
    table_paths = _get_table_paths_for_sheet(zin, sheet_path)
    max_cols = used_cols
    for tp in table_paths:
        try:
            cnt=_read_table_cols_count(zin.read(tp))
            if cnt>max_cols: max_cols=cnt
        except: pass
    new_sheet_xml = _patch_sheet_xml(zin.read(sheet_path), header_row, start_row, max_cols, block_2d)
    last_row = max(header_row, start_row + max(0, len(block_2d)) - 1)
    patched_tables={}
    for tp in table_paths:
        try: patched_tables[tp]=_patch_table_xml(zin.read(tp), header_row, last_row, max_cols)
        except: pass
    out_bio = io.BytesIO()
    with zipfile.ZipFile(out_bio, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            fn=item.filename
            if fn==sheet_path: zout.writestr(item, new_sheet_xml)
            elif fn in patched_tables: zout.writestr(item, patched_tables[fn])
            elif fn.lower()=="[content_types].xml": zout.writestr(item, _strip_calcchain_override(zin.read(fn)))
            elif fn.lower()=="xl/calcchain.xml": continue
            else: zout.writestr(item, zin.read(fn))
    zin.close(); out_bio.seek(0); return out_bio.getvalue()

# ── UI ───────────────────────────────────────────────────────────────
st.title("🎯 Masterfile Automation – Target")
st.caption("Ultra-fast XML writer. Preserves all sheets, styles, formulas, macros.")

st.markdown("<div class='section'><span class='badge badge-target'>Target Marketplace</span> "
            "<span class='badge badge-info'>Template-only writer</span> "
            "<span class='badge badge-ok'>Fast XML Processing</span></div>", unsafe_allow_html=True)

st.markdown("<div class='section'>", unsafe_allow_html=True)
st.markdown("### 📤 Upload Files")

c1, c2 = st.columns([1, 1])
with c1:
    masterfile_file = st.file_uploader("📄 Masterfile Template (.xlsx / .xlsm)", type=["xlsx","xlsm"], help="Upload your Target masterfile template")
with c2:
    onboarding_file = st.file_uploader("🧾 Onboarding Sheet (.xlsx)", type=["xlsx"], help="Upload the onboarding data")

# ── FILTER UI (FIRST COLUMN ONLY) ────────────────────────────────────
st.markdown("#### 🔎 Row filter (by category)")
if onboarding_file is not None:
    try:
        xl = pd.ExcelFile(onboarding_file)
        preview = xl.parse(xl.sheet_names[0], header=0, dtype=str, nrows=200).fillna("")
        preview.columns = [str(c).strip() for c in preview.columns]
        if len(preview.columns) > 0:
            first_col = preview.columns[0]
            st.caption(f"Using onboarding column **{first_col}**")
            vals = sorted({str(v).strip() for v in preview[first_col].astype(str) if str(v).strip() not in ("","nan","none")})
            defaults = []
            if masterfile_file is not None:
                fname = (masterfile_file.name or "").lower()
                defaults = [v for v in vals if v.lower() in fname]
            chosen = st.multiselect("Include only these values from column A", options=vals, default=defaults)
            st.session_state.cat_col = first_col
            st.session_state.cat_values = chosen
        else:
            st.info("No columns found in the first sheet of onboarding.")
    except Exception:
        st.info("Upload a valid onboarding file to enable category filtering.")
else:
    st.info("Upload the onboarding file to enable category filtering.")

# ── Mapping JSON / Output UI ─────────────────────────────────────────
st.markdown("#### 🔗 Mapping JSON")
st.caption("Define how onboarding columns map to masterfile headers")

tab1, tab2 = st.tabs(["📝 Paste JSON", "📁 Upload JSON"])
mapping_json_text, mapping_json_file = "", None
with tab1:
    mapping_json_text = st.text_area(
        "Paste mapping JSON", height=200,
        placeholder='{\n  "TCIN": ["tcin", "item_id"],\n  "Product Title": ["title", "product_name"],\n  "Product Description": ["description", "desc"]\n}',
        help="Map master headers to possible onboarding column names"
    )
with tab2:
    mapping_json_file = st.file_uploader("Or upload mapping.json", type=["json"], key="mapping_file", help="Upload a JSON file with column mappings")

st.markdown("#### 📝 Output Settings")
final_name_input = st.text_input("Final file name (without extension)", value="target_final_masterfile", help="Extension added automatically", max_chars=100)
st.markdown("</div>", unsafe_allow_html=True)

st.divider()
go = st.button("🚀 Generate Final Masterfile", type="primary", use_container_width=True)

# ── Processing ───────────────────────────────────────────────────────
if go:
    overall_t0 = time.time()
    st.markdown("<div class='section'>", unsafe_allow_html=True)
    st.markdown("### 📝 Processing Log")
    log_container = st.empty(); progress_bar = st.progress(0)
    def slog(msg, progress=None):
        log_container.markdown(msg)
        if progress is not None: progress_bar.progress(progress)

    if not masterfile_file or not onboarding_file:
        st.error("❌ Please upload both **Masterfile Template** and **Onboarding** files.")
        st.markdown("</div>", unsafe_allow_html=True); st.stop()
    try:
        ext = (Path(masterfile_file.name).suffix or ".xlsx").lower()
        mime_map = {".xlsx":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsm":"application/vnd.ms-excel.sheet.macroEnabled.12"}

        # Step 1: mapping JSON
        slog("⏳ **Step 1/6:** Parsing mapping JSON...", 0.1)
        try:
            if mapping_json_text.strip(): mapping_raw = json.loads(mapping_json_text)
            elif mapping_json_file: mapping_raw = json.load(mapping_json_file)
            else: st.error("❌ Please provide mapping JSON (paste or upload)."); st.markdown("</div>", unsafe_allow_html=True); st.stop()
        except json.JSONDecodeError as e:
            st.error(f"❌ Invalid JSON format: {e}"); st.markdown("</div>", unsafe_allow_html=True); st.stop()
        if not isinstance(mapping_raw, dict):
            st.error('❌ Mapping JSON must be an object: {"Master header": [aliases...]}'); st.markdown("</div>", unsafe_allow_html=True); st.stop()
        mapping_aliases = {}
        for k, v in mapping_raw.items():
            aliases = v[:] if isinstance(v, list) else [v]
            if k not in aliases: aliases.append(k)
            mapping_aliases[norm(k)] = aliases
        slog(f"✅ Loaded {len(mapping_aliases)} header mappings", 0.2)

        # Step 2: template headers
        slog("⏳ **Step 2/6:** Reading template headers...", 0.3)
        masterfile_file.seek(0); master_bytes = masterfile_file.read()
        wb_ro = load_workbook(io.BytesIO(master_bytes), read_only=True, data_only=True, keep_links=True)
        if MASTER_TEMPLATE_SHEET not in wb_ro.sheetnames:
            st.error(f"❌ Sheet **'{MASTER_TEMPLATE_SHEET}'** not found. Available: {', '.join(wb_ro.sheetnames)}"); st.markdown("</div>", unsafe_allow_html=True); st.stop()
        ws_ro = wb_ro[MASTER_TEMPLATE_SHEET]
        used_cols = worksheet_used_cols(ws_ro, header_rows=(MASTER_DISPLAY_ROW, MASTER_SECONDARY_ROW))
        display_headers   = [ws_ro.cell(row=MASTER_DISPLAY_ROW,   column=c).value or "" for c in range(1, used_cols+1)]
        secondary_headers = [ws_ro.cell(row=MASTER_SECONDARY_ROW, column=c).value or "" for c in range(1, used_cols+1)]
        wb_ro.close()
        slog(f"✅ Loaded {used_cols} template columns", 0.4)

        # Step 3: read onboarding (best sheet)
        slog("⏳ **Step 3/6:** Analyzing onboarding sheet...", 0.5)
        best_xl = pd.ExcelFile(onboarding_file)
        best, best_score, best_info = None, -1, ""
        for sheet in best_xl.sheet_names:
            try:
                df = best_xl.parse(sheet_name=sheet, header=0, dtype=str).fillna("")
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue
            header_set = {norm(c) for c in df.columns}
            matches = sum(any(norm(a) in header_set for a in aliases) for aliases in mapping_aliases.values())
            rows = nonempty_rows(df); score = matches + (0.01 if rows>0 else 0.0)
            if score > best_score:
                best, best_score = (df, sheet), score
                best_info = f"{matches} matched headers, {rows} non-empty rows"
        if best is None: st.error("❌ No readable onboarding sheet found with matching headers."); st.markdown("</div>", unsafe_allow_html=True); st.stop()

        on_df = best[0].fillna("")
        on_df.columns = [str(c).strip() for c in on_df.columns]
        on_headers = list(on_df.columns)
        st.success(f"✅ Using onboarding sheet: **{best[1]}** ({best_info})")

        # Step 3.5: apply category filter (FIRST COLUMN ONLY)
        cat_col = st.session_state.get("cat_col")
        cat_vals = st.session_state.get("cat_values")
        if cat_col and cat_col in on_df.columns and cat_vals:
            before = len(on_df)
            on_df = on_df[on_df[cat_col].astype(str).str.strip().isin(cat_vals)].copy()
            st.info(f"Filtering on **{cat_col}** ∈ {cat_vals} → kept {len(on_df)}/{before} rows.")
        elif cat_col and cat_col in on_df.columns:
            st.warning("No category values selected — no filtering applied.")
        else:
            st.warning("No category column detected — no filtering applied.")

        # Step 3.7: infer from SEO columns (adds new derived columns)
        try:
            seo_cols = select_seo_columns(on_df)
            ordered = order_seo_columns(seo_cols)

            on_df["Gender"] = on_df.apply(lambda r: infer_gender_from_columns(r, ordered), axis=1)
            on_df["health and beauty subtype*"] = on_df.apply(lambda r: infer_hb_subtype_from_columns(r, ordered), axis=1)
            on_df["Health Application*"] = on_df.apply(lambda r: infer_health_app_from_columns(r, ordered), axis=1)
            on_df["Targeted Audience*"] = on_df.apply(lambda r: infer_targeted_audience(r, ordered, r.get("Gender","")), axis=1)
            on_df["Legally Required Information*"] = "Healthcare Disclaimer"
            on_df["Product Form*"] = on_df.apply(lambda r: infer_product_form_from_columns(r, ordered), axis=1)
            on_df["primary flavors"] = on_df.apply(lambda r: infer_primary_flavors_from_columns(r, ordered, 3), axis=1)
            on_df["food and drink form 1"] = on_df.apply(lambda r: infer_food_and_drink_form1_from_columns(r, ordered), axis=1)
            on_df["Prop 65"] = "No"
            on_df["Tax*"] = on_df.apply(lambda r: infer_tax_from_columns(r, ordered), axis=1)

        except Exception:
            on_df["Gender"] = "Gender Neutral"
            on_df["health and beauty subtype*"] = ""
            on_df["Health Application*"] = ""
            on_df["Targeted Audience*"] = "Adult"
            on_df["Legally Required Information*"] = "Healthcare Disclaimer"
            on_df["Product Form*"] = ""
            on_df["primary flavors"] = ""
            on_df["food and drink form 1"] = ""
            on_df["Prop 65"] = "No"
            on_df["Tax*"] = ""

        # refresh headers so mapping sees the new columns
        on_headers = list(on_df.columns)

        # ── Helper attributes we will highlight if empty ──────────────
        HELPER_HEADERS = [
            "health and beauty subtype*",
            "Health Application*",
            "Targeted Audience*",
            "Legally Required Information*",
            "Product Form*",
            "primary flavors",
            "Prop 65",
            "Tax*",
            "food and drink form 1",
        ]
        HELPER_HEADERS_NORM = {norm(h) for h in HELPER_HEADERS}

        # Step 4: mapping
        slog("⏳ **Step 4/6:** Mapping columns...", 0.6)
        series_by_alias = {norm(h): on_df[h] for h in on_headers}
        report_lines = ["#### 🔎 Column Mapping Results"]
        master_to_source = {}; matched_count=0; unmatched_count=0
        helper_cols_idx = set()  # 1-based indices in the template for highlighting

        for c, (disp, sec) in enumerate(zip(display_headers, secondary_headers), start=1):
            eff = disp; eff_norm = norm(eff)
            if not eff_norm: continue

            # record helper column positions by template header
            if eff_norm in HELPER_HEADERS_NORM:
                helper_cols_idx.add(c)

            aliases = mapping_aliases.get(eff_norm, [eff])
            resolved=None; matched_alias=None
            for a in aliases:
                s = series_by_alias.get(norm(a))
                if s is not None: resolved=s; matched_alias=a; break
            if resolved is not None:
                master_to_source[c]=resolved; matched_count+=1
                report_lines.append(f"- ✅ **{eff}** ← `{matched_alias}`")
            else:
                sugg = top_matches(eff, on_headers, 3)
                sug_txt = ", ".join(f"`{name}` ({round(sc*100,1)}%)" for sc,name in sugg) if sugg else "*none*"
                report_lines.append(f"- ❌ **{eff}** ← _no match_. Suggestions: {sug_txt}"); unmatched_count+=1

        st.markdown("\n".join(report_lines))
        st.info(f"📊 Mapping Stats: **{matched_count} matched**, **{unmatched_count} unmatched** out of {len(display_headers)} total columns")

        # Step 5: build data block
        slog("⏳ **Step 5/6:** Building data block...", 0.7)
        n_rows = len(on_df)
        block = [[""] * used_cols for _ in range(n_rows)]
        for col, src in master_to_source.items():
            vals = src.astype(str).tolist(); m=min(len(vals), n_rows)
            for i in range(m):
                v = sanitize_xml_text(vals[i].strip())
                if v and v.lower() not in ("nan","none",""):
                    block[i][col-1] = v
        slog(f"✅ Built data block: {n_rows} rows × {used_cols} columns", 0.8)

        # Step 6: write file (fast XML)
        slog("⏳ **Step 6/6:** Writing final masterfile via fast XML...", 0.85)
        out_bytes = fast_patch_template(master_bytes=master_bytes, sheet_name=MASTER_TEMPLATE_SHEET,
                                        header_row=MASTER_DISPLAY_ROW, start_row=MASTER_DATA_START_ROW,
                                        used_cols=used_cols, block_2d=block)

        # Step 6b: post-highlight empty helper attrs in yellow
        slog("🎨 Applying yellow highlight to empty helper attributes…", 0.92)
        wb = load_workbook(io.BytesIO(out_bytes), keep_vba=(ext == ".xlsm"))
        ws = wb[MASTER_TEMPLATE_SHEET]
        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")

        # for each data row, if helper column value empty, fill yellow
        for i in range(n_rows):
            excel_row = MASTER_DATA_START_ROW + i
            # safety: only highlight existing helper columns within used_cols
            for col_idx in helper_cols_idx:
                if col_idx < 1 or col_idx > used_cols:
                    continue
                val = block[i][col_idx - 1] if i < len(block) and (col_idx - 1) < len(block[i]) else ""
                if str(val).strip() == "":
                    ws.cell(row=excel_row, column=col_idx).fill = yellow

        out_bio2 = io.BytesIO()
        wb.save(out_bio2)
        out_bio2.seek(0)
        out_bytes_final = out_bio2.getvalue()

        st.success("🎉 **Complete!** (with highlight on empty helper attributes)")
        final_base = safe_filename(final_name_input, fallback="target_final_masterfile")
        final_filename = f"{final_base}{ext}"
        st.download_button("⬇️ Download Final Masterfile", data=out_bytes_final, file_name=final_filename,
                           mime=mime_map.get(ext, mime_map[".xlsx"]), key="dl_final_fast", use_container_width=True)

        # Summary metrics
        st.markdown("---")
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("Total Rows", f"{n_rows:,}")
        with c2: st.metric("Total Columns", f"{used_cols}")
        with c3: st.metric("Matched", f"{matched_count}")
        with c4: st.metric("Processing Time", f"{time.time()-overall_t0:.2f}s")

    except Exception as e:
        st.error(f"❌ **Error:** {str(e)}")
        with st.expander("🐛 See full error details"): st.exception(e)
    finally:
        st.markdown("</div>", unsafe_allow_html=True)

